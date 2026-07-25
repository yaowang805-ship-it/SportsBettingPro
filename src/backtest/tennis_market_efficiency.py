"""网球 Pinnacle 市场效率分析 — 多年度、按赛事级别、赔率范围。

数据源: tennis-data.co.uk (ATP/WTA, Pinnacle closing odds)
输出: data/odds/tennis_market_efficiency.json
"""
import io
import json
import sys
import urllib.request
from collections import defaultdict
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent.parent
sys.path.insert(0, str(ROOT))

OUTPUT = ROOT / "data" / "odds" / "tennis_market_efficiency.json"


def download_tennis_xlsx(year: int) -> list:
    """Download tennis Excel file, parse with openpyxl, return rows."""
    try:
        import openpyxl
    except ImportError:
        print("  openpyxl not available")
        return []

    url = f"http://www.tennis-data.co.uk/{year}/{year}.xlsx"
    try:
        req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
        with urllib.request.urlopen(req, timeout=30) as resp:
            raw = resp.read()
        wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True)
        ws = wb[wb.sheetnames[0]]

        headers = [str(c.value) for c in next(ws.iter_rows(min_row=1, max_row=1))]
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            rows.append(dict(zip(headers, row)))
        return rows
    except Exception as e:
        print(f"  {year}: {e}")
        return []


def analyze_tennis(rows: list) -> dict:
    """Analyze Pinnacle efficiency by series level, odds range, surface."""
    results = {
        "total_matches": 0,
        "total_bets": 0,
        "total_stake": 0.0,
        "total_profit": 0.0,
        "total_won": 0,
        "by_series": defaultdict(lambda: {"bets": 0, "stake": 0.0, "profit": 0.0, "won": 0,
                                            "by_odds": defaultdict(lambda: {"bets": 0, "stake": 0, "profit": 0, "won": 0})}),
        "by_odds_range": defaultdict(lambda: {"bets": 0, "stake": 0.0, "profit": 0.0, "won": 0}),
        "flb_analysis": defaultdict(lambda: {"bets": 0, "stake": 0, "profit": 0, "won": 0}),
    }

    for r in rows:
        try:
            psw = float(r.get("PSW", 0) or 0)
            psl = float(r.get("PSL", 0) or 0)
            if not (psw > 1 and psl > 1):
                continue
            comment = str(r.get("Comment", "") or "")
            if "Retired" in comment or "Walkover" in comment:
                continue
        except (ValueError, TypeError):
            continue

        series = str(r.get("Series", "Unknown"))
        surface = str(r.get("Surface", "Unknown"))

        results["total_matches"] += 1

        for odds, hit in [(psw, True), (psl, False)]:
            stake = 1.0
            results["total_bets"] += 1
            results["total_stake"] += stake
            if hit:
                results["total_won"] += 1
                results["total_profit"] += odds - 1
            else:
                results["total_profit"] -= 1

            # By odds range
            if odds < 1.3: bucket = "<1.3"
            elif odds < 1.5: bucket = "1.3-1.5"
            elif odds < 2.0: bucket = "1.5-2.0"
            elif odds < 3.0: bucket = "2.0-3.0"
            elif odds < 5.0: bucket = "3.0-5.0"
            elif odds < 10.0: bucket = "5.0-10.0"
            else: bucket = ">10.0"

            br = results["by_odds_range"][bucket]
            br["bets"] += 1; br["stake"] += stake
            if hit: br["won"] += 1; br["profit"] += odds - 1
            else: br["profit"] -= 1

            # By series + odds
            sr = results["by_series"][series]
            sr["bets"] += 1; sr["stake"] += stake
            if hit: sr["won"] += 1; sr["profit"] += odds - 1
            else: sr["profit"] -= 1
            so = sr["by_odds"][bucket]
            so["bets"] += 1; so["stake"] += stake
            if hit: so["won"] += 1; so["profit"] += odds - 1
            else: so["profit"] -= 1

            # FLB: bet only low odds (<2.0)
            if odds < 2.0:
                flb = results["flb_analysis"]["low_odds"]
            elif odds < 5.0:
                flb = results["flb_analysis"]["mid_odds"]
            else:
                flb = results["flb_analysis"]["high_odds"]
            flb["bets"] += 1; flb["stake"] += stake
            if hit: flb["won"] += 1; flb["profit"] += odds - 1
            else: flb["profit"] -= 1

    return results


def generate_recommendations(results: dict) -> dict:
    """Generate data-driven recommendations from analysis results."""
    recs = {
        "pinnacle_vig": round(-results["total_profit"] / results["total_stake"] * 100, 2),
        "total_matches": results["total_matches"],
        "total_bets": results["total_bets"],
        "by_series": {},
        "flb_summary": {},
        "recommended_weights": {},
        "recommended_odds_limits": {},
    }

    # FLB summary
    for bucket in ["low_odds", "mid_odds", "high_odds"]:
        d = results["flb_analysis"].get(bucket, {})
        if d["bets"] > 0:
            recs["flb_summary"][bucket] = {
                "bets": d["bets"],
                "roi": round(d["profit"] / d["stake"] * 100, 2),
                "win_rate": round(d["won"] / d["bets"] * 100, 1),
            }

    # By series
    for series, sr in sorted(results["by_series"].items(), key=lambda x: -x[1]["bets"]):
        if sr["bets"] < 10:
            continue
        series_rec = {
            "bets": sr["bets"],
            "roi": round(sr["profit"] / sr["stake"] * 100, 2),
            "win_rate": round(sr["won"] / sr["bets"] * 100, 1),
            "vig": round(-sr["profit"] / sr["stake"] * 100, 2),
            "by_odds": {},
        }
        # Odds analysis per series
        for bucket in ["<1.3", "1.3-1.5", "1.5-2.0", "2.0-3.0", "3.0-5.0", "5.0-10.0", ">10.0"]:
            so = sr["by_odds"].get(bucket, {})
            if so.get("bets", 0) > 5:
                series_rec["by_odds"][bucket] = {
                    "bets": so["bets"],
                    "roi": round(so["profit"] / so["stake"] * 100, 2) if so["stake"] else 0,
                    "win_rate": round(so["won"] / so["bets"] * 100, 1) if so["bets"] else 0,
                }
        recs["by_series"][series] = series_rec

        # Recommended weight: inverse vig, scaled to football baseline
        if series_rec["vig"] > 0:
            weight = round(4.0 / series_rec["vig"], 2)  # 4% = football avg vig
            recs["recommended_weights"][series] = min(1.5, max(0.3, weight))

        # Recommended odds limit: where ROI drops below -5%
        max_safe_odds = 10.0
        for bucket in ["<1.3", "1.3-1.5", "1.5-2.0", "2.0-3.0", "3.0-5.0", "5.0-10.0", ">10.0"]:
            so = sr["by_odds"].get(bucket, {})
            if so.get("bets", 0) > 5 and so.get("roi", 0) > -5:
                # Still profitable (within vig), extend limit
                pass
            elif so.get("bets", 0) > 10 and so.get("roi", 0) < -8:
                # Significant FLB, set limit here
                limits = {"<1.3": 1.3, "1.3-1.5": 1.5, "1.5-2.0": 2.0,
                          "2.0-3.0": 3.0, "3.0-5.0": 5.0, "5.0-10.0": 10.0}
                max_safe_odds = min(max_safe_odds, limits.get(bucket, 10.0))
        recs["recommended_odds_limits"][series] = max_safe_odds

    return recs


if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument("--years", type=int, default=3, help="Years to download")
    args = parser.parse_args()

    current_year = 2025
    all_rows = []
    for y in range(current_year - args.years, current_year + 1):
        print(f"Downloading {y}...", end=" ", flush=True)
        rows = download_tennis_xlsx(y)
        print(f"{len(rows)} matches")
        all_rows.extend(rows)

    print(f"\nTotal: {len(all_rows)} matches")
    results = analyze_tennis(all_rows)
    recs = generate_recommendations(results)

    t = results
    vig = -t["total_profit"] / t["total_stake"] * 100
    print(f"\n=== 🎾 网球 Pinnacle 效率 ({t['total_matches']:,}场, {t['total_bets']:,}笔) ===")
    print(f"Pinnacle抽水: {vig:.2f}%")
    print(f"胜率: {t['total_won']/t['total_bets']*100:.1f}%")

    print(f"\nFLB分析:")
    for bucket, d in recs["flb_summary"].items():
        print(f"  {bucket}: {d['bets']:,}笔, ROI={d['roi']:+.2f}%, 胜率={d['win_rate']}%")

    print(f"\n按赛事级别:")
    for series, sr in sorted(recs["by_series"].items(), key=lambda x: -x[1]["bets"]):
        print(f"  {series}: {sr['bets']:,}笔, vig={sr['vig']:.2f}%, 建议权重={recs['recommended_weights'].get(series, 1.0):.2f}")
        for bucket, so in sr.get("by_odds", {}).items():
            print(f"    {bucket}: {so['bets']:,}笔, ROI={so['roi']:+.1f}%, 胜率={so['win_rate']}%")

    # Save
    output_data = {"results": recs, "raw_summary": {
        "total_matches": t["total_matches"],
        "total_bets": t["total_bets"],
        "total_vig": round(vig, 2),
    }}
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    json.dump(output_data, open(OUTPUT, "w"), ensure_ascii=False, indent=2, default=str)
    print(f"\n✅ 保存到 {OUTPUT}")
